from openpyxl import load_workbook
import os
from pprint import pprint
import re
import json

path = 'I:\\My Drive\\Academic\\Attendance Sheets\\'


def open_workbook(filename):
    file_path = path + filename
    values = []
    try:
        attendance_book = load_workbook(file_path, read_only=False)
    except Exception as e:
        print(e)
        return values

    # print('opening file ' + filename)
    sheet_name = "Grad Dip(Learning)"
    if sheet_name not in attendance_book.sheetnames:
        print(f"'{sheet_name}' not found. Quitting.")
        return values
    sheet = attendance_book[sheet_name]
    for i in range(10, len(sheet['A'])):
        values.append(sheet['A'][i].value)

    return values


def save_student_ids():
    files = get_all_files(path)
    student_id_pattern = r"([a-zA-Z]{2}[a-zA-Z]?\d{6})"
    ids = []
    for file in files:
        ids += open_workbook(file)
    ids = list(set(ids))
    ids_str = ' '.join(map(str, ids))
    valid_ids = re.findall(pattern=student_id_pattern, string=ids_str)
    # Serializing json
    json_object = json.dumps({"student_ids": valid_ids}, indent=3)

    # Writing to sample.json
    with open("student_ids.json", "w") as outfile:
        outfile.write(json_object)


def print_student_ids():
    # Opening JSON file
    with open('student_ids.json', 'r') as openfile:
        # Reading from json file
        json_object = json.load(openfile)

    print(type(json_object))
    pprint(json_object)


def get_all_files(dir_path):
    return os.listdir(dir_path)


if __name__ == "__main__":
    print_student_ids()
